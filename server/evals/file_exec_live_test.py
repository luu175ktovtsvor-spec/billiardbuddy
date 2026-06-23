# -*- coding: utf-8 -*-
"""真实架构·【工具执行】测试：让 AI 在完整壳子里【改文件 + 跑命令】，抓改前/改后真实结果。

和 architecture_live_test 一样走真 app/真循环/真工具/真 MiMo；区别是这次考的是【执行类工具】：
edit_excel / edit_file / write_file / run_command。物料(Excel/文案/名单)提前造好，作为"老板当场选定的文件"
(selected_files → ctx.allowed_paths)交给 AI 改。permission_mode=full + 关闭自动放行上限 → 审批闸自动放行、工具内联执行。

指令分三档考它：清晰 / 中等(需用行业知识) / 模糊(理应先问澄清，别瞎改)，外加一条跑命令。
每条记录【改前内容 → AI 走了哪些工具 → 改后内容 → 对照标准合不合格】。
跑法：MIMO_KEY='sk-...' uv run python evals/file_exec_live_test.py
"""
import os
import sys
import asyncio
import json
import uuid
import logging
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
import tempfile
os.environ["DATABASE_URL"] = f"sqlite+aiosqlite:///{tempfile.mktemp(suffix='.db')}"
os.environ["DESKTOP_LOCAL"] = "1"
os.environ.setdefault("SECRET_KEY", "fileexec-secret")
os.environ.setdefault("BYOK_ENCRYPT_KEY", "aGVsbG93b3JsZGhlbGxvd29ybGRoZWxsb3dvcmxkMTI=")
os.environ.setdefault("UPLOAD_DIR", tempfile.mkdtemp())

for _n in ("httpx", "httpcore", "openai", "openai._base_client"):
    logging.getLogger(_n).setLevel(logging.WARNING)

import httpx  # noqa: E402
from httpx import ASGITransport  # noqa: E402
import openpyxl  # noqa: E402

from main import app  # noqa: E402
from db.session import async_session  # noqa: E402
from db.init_local import init_local_db  # noqa: E402
from models.user import User  # noqa: E402
from models.store import Store, StoreMember  # noqa: E402
from api.deps import get_current_user, get_current_store  # noqa: E402
from services.ai import factory  # noqa: E402
from services.ai.providers.deepseek import DeepSeekProvider  # noqa: E402

MIMO = DeepSeekProvider(api_key=os.environ["MIMO_KEY"], base_url="https://api.xiaomimimo.com/v1",
                        default_model="mimo-v2.5", timeout=300.0)
factory.ProviderFactory.get_text_provider_for_store = classmethod(lambda cls, store: MIMO)
factory.ProviderFactory.get_orchestration_provider = classmethod(lambda cls: MIMO)

WORK = ROOT.parent / "测试项目的AI Agent功能" / "文件执行测试"
XLSX = WORK / "营业额报表.xlsx"
COPY = WORK / "朋友圈草稿.md"

OVER_OFFER = "🎉劲爆充值！充1万送1万！全城最低价！台费终身免费畅打！名额有限速来！"


def make_xlsx(jun=12, may=9):
    WORK.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "营业额"
    ws.append(["月份", "营业额(万)"])
    for m, v in [("1月", 10), ("2月", 11), ("3月", 13), ("4月", 12), ("5月", may), ("6月", jun)]:
        ws.append([m, v])
    wb.save(XLSX)


def read_xlsx():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    return {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}


def make_copy():
    WORK.mkdir(parents=True, exist_ok=True)
    COPY.write_text(f"# 周末充值活动朋友圈\n\n各位球友！这个周末搞大活动：\n{OVER_OFFER}\n\n地址老地方，来玩呀～\n",
                    encoding="utf-8")


# (标签, 指令档位, 指令, 准备函数, 选定文件, 标准说明)
def TASKS():
    return [
        ("01_清晰改Excel", "清晰", f"把这个营业额报表里 6月的营业额改成 15万。文件：{XLSX}",
         lambda: make_xlsx(jun=12), [str(XLSX)],
         "标准：6月单元格应=15。读改后xlsx核对。"),
        ("02_中等改文案需用知识", "中等",
         f"这条朋友圈草稿里的充值活动力度太猛了，按咱们行业规矩调合理点。文件：{COPY}",
         make_copy, [str(COPY)],
         "标准：'充1万送1万'降到≤10%力度、'全城最低/终身免费'违规词去掉。读改后md核对。"),
        ("03_模糊改报表", "模糊", f"这报表里好像有个数不太对，你帮我看看、改一下。文件：{XLSX}",
         lambda: make_xlsx(jun=12, may=90), [str(XLSX)],
         "标准(对的行为)：要么先问'哪个数/按什么改'，要么识别出5月=90是异常(邻月~10)并修；【不能】乱改一个正常数。"),
        ("04_清晰跑命令", "清晰", f"用命令看一下这个文件夹（{WORK}）里都有哪些文件，列出来。",
         lambda: None, [],
         "标准：调 run_command（如 ls）真执行并列出文件名。"),
    ]


async def seed():
    await init_local_db()
    async with async_session() as db:
        u = User(id=uuid.uuid4(), phone="13800000000", password_hash="x", name="测试老板")
        db.add(u)
        await db.flush()
        s = Store(id=uuid.uuid4(), owner_id=u.id, name="星光台球俱乐部", city="成都")
        # 关闭自动放行上限 → full 模式下文件/命令自动放行、内联执行
        if hasattr(s, "agent_auto_spend_limit"):
            s.agent_auto_spend_limit = -1
        db.add(s)
        await db.flush()
        db.add(StoreMember(store_id=s.id, user_id=u.id, role="owner"))
        await db.commit()
        return u, s


async def run_task(client, store, msg, files):
    tools, final, approvals, errors = [], "", [], []
    body = {"message": msg, "permission_mode": "full", "full_disk_access": True}
    if files:
        body["selected_files"] = files
    try:
        async with client.stream("POST", "/api/v1/agent/chat", json=body,
                                 timeout=300.0) as resp:
            async for line in resp.aiter_lines():
                if not line.startswith("data:"):
                    continue
                try:
                    ev = json.loads(line[5:].strip())
                except Exception:
                    continue
                t = ev.get("type")
                if t == "tool_call":
                    tools.append(ev.get("tool"))
                elif t == "approval_request":
                    approvals.append(ev.get("tool"))
                elif t == "final":
                    final = ev.get("content") or ""
                elif t == "error":
                    errors.append(ev.get("error") or ev.get("message") or str(ev))
    except Exception as e:
        errors.append(f"(连接异常) {e!r}")
    return {"tools": tools, "final": final, "approvals": approvals, "errors": errors}


def snapshot(files):
    """改前/改后内容快照(给报告对照用)。"""
    snap = {}
    if str(XLSX) in files or not files:
        if XLSX.exists():
            snap["营业额报表"] = read_xlsx()
    if str(COPY) in files:
        if COPY.exists():
            snap["朋友圈草稿"] = COPY.read_text(encoding="utf-8")
    return snap


async def main():
    WORK.mkdir(parents=True, exist_ok=True)
    u, s = await seed()
    app.dependency_overrides[get_current_user] = lambda: u
    app.dependency_overrides[get_current_store] = lambda: s
    transport = ASGITransport(app=app)
    report = ["# 工具执行测试 · 改文件/跑命令（真实壳子端到端）\n",
              "> 物料提前造好交给 AI（selected_files）；permission_mode=full 自动放行、工具内联执行。",
              "> 指令分【清晰/中等/模糊】三档 + 一条跑命令。每条：改前 → AI 调了哪些工具 → 改后 → 对照标准。\n"]
    async with httpx.AsyncClient(transport=transport, base_url="http://exec") as client:
        for label, level, msg, prep, files, std in TASKS():
            prep()
            before = snapshot(files)
            print(f"▶ {label}（{level}）跑中...", flush=True)
            r = await run_task(client, s, msg, files)
            after = snapshot(files)
            changed = before != after
            print(f"  工具={r['tools'] or '无'} | 审批卡={r['approvals'] or '无'} | 文件变了={changed} | 错={r['errors'] or '无'}", flush=True)
            report.append(f"\n## {label}（指令档位：{level}）\n")
            report.append(f"**指令：** {msg}\n")
            report.append(f"**{std}**\n")
            report.append(f"**AI 实际调用工具：** {r['tools'] or '（没调工具）'}　|　审批卡：{r['approvals'] or '无'}　|　报错：{r['errors'] or '无'}\n")
            if before:
                report.append(f"**改前：**\n```\n{json.dumps(before, ensure_ascii=False, indent=2)}\n```\n")
            if after:
                report.append(f"**改后：**\n```\n{json.dumps(after, ensure_ascii=False, indent=2)}\n```\n")
            report.append(f"**文件是否被修改：** {'✅ 是' if changed else '⚠️ 否（没改/或先问澄清）'}\n")
            report.append(f"**AI 收尾回复：** {r['final'][:300]}\n")
    out = WORK / "00_执行测试结果.md"
    out.write_text("\n".join(report), encoding="utf-8")
    print(f"\n===== 完成。结果：{out} =====")


if __name__ == "__main__":
    asyncio.run(main())
