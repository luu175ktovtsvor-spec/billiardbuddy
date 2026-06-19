# -*- coding: utf-8 -*-
"""真 LLM 端到端验证：接 MiMo，建真文件夹，让大模型【自己用每个文件/命令工具】，
看它会不会正确选用工具、真把事办成（不是单元测试调函数，是真模型决策）。

- key 走环境变量 MIMO_KEY（脚本里无 key，可入库）。
- 不连 DB：run_agent_loop 注入 provider + ctx 即可（文件工具只用 ctx）。
- 模拟 L3「完全访问模式」：permission_mode=full + full_disk_access=True + auto_spend_limit=-1
  （自己动手、不逐个弹卡——非交互测试需要它真执行）。
跑法：  MIMO_KEY='sk-...' uv run python evals/agent_tools_e2e.py
"""
import os
import sys
import asyncio
import tempfile
import shutil
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ.setdefault("DESKTOP_LOCAL", "1")  # 让系统提示带上文件/完全访问 hint

from services.ai.providers.deepseek import DeepSeekProvider  # noqa: E402
from services.ai.base import TextRequest  # noqa: E402
from services.agent.registry import ToolRegistry  # noqa: E402
from services.agent.local_tools import register_local_tools  # noqa: E402
from services.agent.context import AgentContext  # noqa: E402
from services.agent.loop import run_agent_loop  # noqa: E402
from api.v1.agent import compose_agent_system_prompt  # noqa: E402

MODEL = "mimo-v2.5"
BASE = "https://api.xiaomimimo.com/v1"


def _provider():
    return DeepSeekProvider(api_key=os.environ["MIMO_KEY"], base_url=BASE, default_model=MODEL)


def _setup_files(folder: Path):
    """每个用例前重置标准测试文件，保证用例独立、验证确定。"""
    for p in list(folder.iterdir()):
        shutil.rmtree(p) if p.is_dir() else p.unlink()
    (folder / "notes.txt").write_text("周末活动方案：充500送100，办个比赛聚人气", encoding="utf-8")
    (folder / "readme.md").write_text("# 我的台球房\n联系电话 123456", encoding="utf-8")
    sub = folder / "sub"
    sub.mkdir()
    (sub / "data.txt").write_text("本月会员数 88 人", encoding="utf-8")


def _ctx(folder: Path):
    return AgentContext(full_disk_access=True, permission_mode="full",
                        auto_spend_limit=-1, allowed_paths=[str(folder)])


def _read(fo: Path, name: str) -> str:
    try:
        return (fo / name).read_text(encoding="utf-8")
    except Exception:
        return ""


async def _run(p, reg, folder, sysp, name, expect, request, verify):
    _setup_files(folder)
    try:
        res = await run_agent_loop(user_message=request, registry=reg, ctx=_ctx(folder),
                                   system_prompt=sysp, provider=p, model=MODEL, max_turns=6)
    except Exception as e:
        print(f"❌ [{name}] 循环抛错: {e!r}")
        return name, False
    called = [s.tool_name for s in res.steps if s.type == "tool_call"]
    try:
        ok, detail = verify(res, folder, called)
    except Exception as e:
        ok, detail = False, f"验证抛错 {e!r}"
    print(f"{'✅' if ok else '❌'} [{name}] 期望≈{expect} | 实际调用={called or '无'} | 办成={ok}（{detail}）")
    print(f"     答复: {(res.final_text or '')[:90]!r}")
    return name, ok


async def main():
    p = _provider()
    r = await p.generate(TextRequest(prompt="只回复：连上了", max_tokens=2000))
    print(f"[连通] MiMo {MODEL} → {(r.content or '')[:20]!r}\n")

    reg = ToolRegistry()
    register_local_tools(reg)
    tool_names = sorted(s["function"]["name"] for s in reg.to_openai_tools())
    print(f"[已装工具] {tool_names}\n")

    sysp = compose_agent_system_prompt("（测试台球房）", "", full_disk=True)
    folder = Path(tempfile.mkdtemp(prefix="agent_e2e_"))
    F = str(folder)
    print(f"[测试文件夹] {F}\n")

    cases = [
        ("列目录 list_files", "list_files",
         f"列一下文件夹 {F} 里都有哪些文件和子文件夹",
         lambda res, fo, c: ("list_files" in c or "find_files" in c or "notes.txt" in (res.final_text or ""), "用了列/找类")),
        ("找文件 find_files", "find_files",
         f"在 {F} 这个文件夹里，把所有 .txt 文件都找出来，告诉我有哪些",
         lambda res, fo, c: ("notes.txt" in (res.final_text or "") and "data.txt" in (res.final_text or ""), "答复含两个txt")),
        ("搜内容 search_in_files", "search_in_files",
         f"{F} 这个文件夹里，哪个文件写了「充500送100」？只告诉我文件名",
         lambda res, fo, c: ("notes.txt" in (res.final_text or ""), "答复指向notes.txt")),
        ("读文件 read_file", "read_file",
         f"读一下 {F}/notes.txt 写了什么，复述要点给我",
         lambda res, fo, c: ("500" in (res.final_text or "") or "比赛" in (res.final_text or ""), "答复含原文要点")),
        ("写文件 write_file", "write_file",
         f"在 {F} 里新建一个文件 plan.txt，内容写：周六晚八点八球比赛，报名费30",
         lambda res, fo, c: ((fo / "plan.txt").exists() and "比赛" in _read(fo, "plan.txt"), f"plan.txt存在={(fo / 'plan.txt').exists()}")),
        ("改文件 edit_file", "edit_file",
         f"把 {F}/notes.txt 里的「充500送100」改成「充1000送300」",
         lambda res, fo, c: ("充1000送300" in _read(fo, "notes.txt") and "充500送100" not in _read(fo, "notes.txt"), "notes.txt已改")),
        ("跑命令 run_command", "run_command",
         f"在 {F} 目录下跑一条 ls 命令，把文件列出来",
         lambda res, fo, c: ("run_command" in c, "确实调了run_command")),
        ("★自主找+改(你的场景)", "list/find→read→edit",
         f"我刚在 {F} 里放了个记台球房周末活动的文本文件，活动力度开太大了，帮我找到它、把送的额度改成原来的一半",
         lambda res, fo, c: ("充500送100" not in _read(fo, "notes.txt") and _read(fo, "notes.txt") != "", "notes.txt力度被改动")),
    ]

    results = []
    for name, expect, req, verify in cases:
        results.append(await _run(p, reg, folder, sysp, name, expect, req, verify))

    # 改表 edit_excel（需 openpyxl）
    try:
        import openpyxl

        def _excel_verify(res, fo, c):
            v = openpyxl.load_workbook(fo / "report.xlsx").active["B2"].value
            return (str(v) == "88000", f"B2={v}")

        _setup_files(folder)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A2"], ws["B2"] = "营业额", 32000
        wb.save(folder / "report.xlsx")
        # 注意：此用例 _run 会重置文件，故单独不走 _setup_files——改为内联跑
        res = await run_agent_loop(
            user_message=f"把 {F}/report.xlsx 的 B2 单元格改成 88000",
            registry=reg, ctx=_ctx(folder), system_prompt=sysp, provider=p, model=MODEL, max_turns=6)
        called = [s.tool_name for s in res.steps if s.type == "tool_call"]
        ok, detail = _excel_verify(res, folder, called)
        print(f"{'✅' if ok else '❌'} [改表 edit_excel] 实际调用={called or '无'} | 办成={ok}（{detail}）")
        results.append(("改表 edit_excel", ok))
    except ImportError:
        print("⚠️ [改表 edit_excel] openpyxl 没装，跳过")

    print("\n===== 汇总(真模型能不能正确用上工具、把事办成) =====")
    for name, ok in results:
        print(f"  {'✅' if ok else '❌'} {name}")
    passed = sum(1 for _, ok in results if ok)
    print(f"\n{passed}/{len(results)} 项通过")
    shutil.rmtree(folder, ignore_errors=True)


if __name__ == "__main__":
    asyncio.run(main())
