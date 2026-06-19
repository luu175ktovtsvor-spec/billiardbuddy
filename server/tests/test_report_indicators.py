"""A-1 报表抽取器测试：用 openpyxl 在 tmp_path 造合成小表（不依赖根目录那张真表，它不在 git）。

断言 extract_report_indicators 算出的 团购占比/台费占比/助教费占比/客单价 与手算一致；
再断言缺列时 best-effort 不崩（返回空指标 + 提示 summary）。
"""
import openpyxl
import pytest

from services.report_reader import extract_report_indicators


def _make_full_table(path):
    """造一张双层表头（row1 大类 / row2 细列）的『收入日记账优化』富表 + 一张支出表。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收入日记账优化"

    # row1 大类（占位，抽取器看 row2 细列）
    headers_r1 = ["日期", "总营收", "总营收", "客单", "费用", "费用", "费用",
                  "会员", "会员", "客流", "团购", "团购占比", "翻台",
                  "金腿", "金腿"]
    # row2 细列（抽取器据此 fuzzy 匹配）
    headers_r2 = ["日期", "总营业额", "总单数", "平均每单价", "商品费", "总开台费", "总助教费",
                  "会员充值", "会员消费储值扣款", "新客人数", "总团购费", "团购费占比", "平均每桌时长",
                  "乔氏金腿开台次数", "乔氏金腿每单平均时长"]
    ws.append(headers_r1)
    ws.append(headers_r2)

    # 3 行逐日已知数字
    # 列: 日期, 营业额, 单数, 客单价, 商品费, 开台费, 助教费, 充值, 扣款, 新客, 团购费, 团购占比%, 翻台h, 金腿开台, 金腿时长
    rows = [
        ["6-1", 10000, 100, 100, 800, 6000, 2000, 5000, 1000, 30, 3000, 30, 4.0, 20, 1.5],
        ["6-2", 20000, 200, 100, 1600, 12000, 4000, 3000, 2000, 40, 5000, 25, 5.0, 30, 2.0],
        ["6-3", 10000, 100, 100, 600, 6000, 2000, 2000, 3000, 30, 2000, 20, 3.0, 10, 1.0],
    ]
    for r in rows:
        ws.append(r)

    # 支出表：房租逐日列
    ws2 = wb.create_sheet("支出日记账")
    ws2.append(["日期", "房租", "水电"])
    ws2.append(["6-1", 4000, 500])
    ws2.append(["6-2", 4000, 500])
    ws2.append(["6-3", 4000, 500])

    wb.save(path)


def test_extract_indicators_matches_manual(tmp_path):
    p = tmp_path / "full.xlsx"
    _make_full_table(str(p))
    ind, summary = extract_report_indicators(str(p))

    # 手算：总营业额 = 10000+20000+10000 = 40000
    assert ind["total_revenue"] == 40000

    # 团购占比：优先用「团购费占比」列均值 = (30+25+20)/3 = 25.0
    assert ind["groupbuy_ratio_pct"] == pytest.approx(25.0)

    # 台费占比 = 总开台费(6000+12000+6000=24000) / 40000 = 60.0%
    assert ind["table_fee_ratio_pct"] == pytest.approx(60.0)

    # 助教费占比 = (2000+4000+2000=8000) / 40000 = 20.0%
    assert ind["assistant_fee_ratio_pct"] == pytest.approx(20.0)

    # 商品费占比 = (800+1600+600=3000) / 40000 = 7.5%
    assert ind["goods_fee_ratio_pct"] == pytest.approx(7.5)

    # 客单价：优先列均值 = (100+100+100)/3 = 100
    assert ind["avg_price"] == pytest.approx(100.0)

    # 翻台率（每桌平均时长）= (4+5+3)/3 = 4.0
    assert ind["turnover_hours"] == pytest.approx(4.0)

    # 会员充值 10000 / 扣款 6000
    assert ind["member_recharge"] == 10000
    assert ind["member_deduct"] == 6000

    # 商品费(3000) < 助教费(8000)
    assert ind["goods_below_assistant"] is True

    # 房租占比 = (4000*3=12000) / 40000 = 30.0%
    assert ind["rent_ratio_pct"] == pytest.approx(30.0)

    # 各区：金腿开台 = 20+30+10 = 60；只有一区 → 占比 100%
    assert "areas" in ind
    assert ind["areas"]["乔氏金腿"]["open"] == 60
    assert ind["areas"]["乔氏金腿"]["open_ratio_pct"] == pytest.approx(100.0)

    # summary 是可读串，含关键标尺词
    assert "团购占比" in summary
    assert "台费" in summary


def test_groupbuy_ratio_fallback_from_fee(tmp_path):
    """没有「团购费占比」列时，团购占比 = 总团购费 / 总营业额。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收入日记账"
    ws.append(["日期", "营收", "费用", "团购"])
    ws.append(["日期", "总营业额", "总开台费", "总团购费"])
    ws.append(["6-1", 10000, 6000, 2000])
    ws.append(["6-2", 10000, 6000, 2000])
    p = tmp_path / "nofratio.xlsx"
    wb.save(str(p))

    ind, _ = extract_report_indicators(str(p))
    # 团购占比 = (2000+2000=4000) / (10000+10000=20000) = 20.0%
    assert ind["groupbuy_ratio_pct"] == pytest.approx(20.0)


def test_missing_columns_best_effort_no_crash(tmp_path):
    """缺一堆列时 best-effort：能算的算、缺的进 missing，绝不抛错。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收入日记账"
    ws.append(["日期", "营收"])
    ws.append(["日期", "总营业额"])
    ws.append(["6-1", 10000])
    ws.append(["6-2", 20000])
    p = tmp_path / "sparse.xlsx"
    wb.save(str(p))

    ind, summary = extract_report_indicators(str(p))
    assert ind["total_revenue"] == 30000
    assert "missing" in ind
    # 团购占比/台费占比/翻台率等都该在缺列里
    assert "团购占比" in ind["missing"]
    assert "台费占比" in ind["missing"]
    assert isinstance(summary, str) and summary


def test_bad_file_degrades_gracefully(tmp_path):
    """坏文件（非 xlsx）不崩，返回空指标 + 提示 summary。"""
    p = tmp_path / "broken.xlsx"
    p.write_text("this is not an excel file", encoding="utf-8")
    ind, summary = extract_report_indicators(str(p))
    assert ind == {}
    assert isinstance(summary, str) and summary


def test_no_rich_sheet(tmp_path):
    """没有富表（没营业额/团购/开台列）时不崩。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "随便一张表"
    ws.append(["甲", "乙"])
    ws.append([1, 2])
    p = tmp_path / "norich.xlsx"
    wb.save(str(p))
    ind, summary = extract_report_indicators(str(p))
    assert isinstance(ind, dict)
    assert isinstance(summary, str) and summary
