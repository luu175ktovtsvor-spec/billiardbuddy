"""真实 report_forms/*.yaml 的渲染 + prompt_key 解析冒烟。

护栏：5 张表的配置必须能被渲染器渲出合法 xlsx，且 narrative.prompt_key 必须
在 PromptEngine 里真实存在（防"schema 引用了不存在的 prompt"的耦合 bug）。
"""
import io

from openpyxl import load_workbook

from services.ai.prompt_engine import get_prompt_engine
from services.report_excel import render_report
from services.report_schema import all_report_schemas

_EXPECTED = {
    "manager_daily", "frontdesk_daily", "coach_main_daily",
    "coach_assistant_daily", "assistant_manager_daily",
}


def _sample_data(schema: dict) -> dict:
    base = {"store_name": "示例球房", "date": "2026-06-14", "weekday": "周日"}
    if schema["shape"] == "roster":
        metric_keys = [c["key"] for c in schema["columns"] if c["key"] != "name"]
        base["rows"] = [
            {"name": "助教A", **{k: 3 for k in metric_keys}},
            {"name": "助教B", **{k: 5 for k in metric_keys}},
        ]
    else:
        for g in schema["groups"]:
            for f in g["fields"]:
                base[f["key"]] = 1
        if schema["shape"] == "personal":
            base["_cumulative"] = {k: 10 for k in schema.get("cumulative_fields", [])}
    return base


def test_all_real_schemas_render_valid_xlsx():
    schemas = all_report_schemas()
    assert set(schemas) >= _EXPECTED
    for key, schema in schemas.items():
        buf = render_report(schema, _sample_data(schema), narrative=f"{key} 叙事样例。")
        wb = load_workbook(io.BytesIO(buf))   # 能打开即合法
        assert wb.active.max_row >= 1, key


def test_all_narrative_prompt_keys_resolve():
    engine = get_prompt_engine()
    for key, schema in all_report_schemas().items():
        pk = schema["narrative"]["prompt_key"]
        assert pk in engine._templates, f"{key} 的 prompt_key {pk} 未找到"
