"""Excel 渲染纯函数单测：传 schema+data+narrative → bytes，openpyxl 读回断言。"""
import io

from openpyxl import load_workbook

from services.report_excel import render_report

_FLAT_SCHEMA = {
    "key": "manager_daily", "shape": "flat", "title": "{store_name}·店长日报·{date}",
    "groups": [{"name": "今日数据", "fields": [
        {"key": "revenue", "label": "营业额", "unit": "元"},
    ]}],
    "narrative": {"outputs": ["今日总结"]},
    "excel": {"layout": "sectioned"},
}


def test_flat_renders_title_and_value():
    buf = render_report(
        _FLAT_SCHEMA,
        data={"store_name": "示例球房", "date": "2026-06-14", "revenue": 5800},
        narrative="今日总结：营业额回升。",
    )
    assert isinstance(buf, (bytes, bytearray)) and len(buf) > 0
    ws = load_workbook(io.BytesIO(buf)).active
    flat = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert any(v == "示例球房·店长日报·2026-06-14" for v in flat)       # 标题
    assert any("5800" in str(v) for v in flat)                          # 数值
    assert any("营业额回升" in str(v) for v in flat)                    # 叙事
