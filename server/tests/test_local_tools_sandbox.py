"""本地文件工具沙箱机制测试（桌面版"报表直改"的安全核心）。

沙箱 = 内容库 + 用户当场经文件选择器选定的文件/目录。验证：
- 库内放行；库外未选文件拒绝；用户选定的库外文件放行；路径穿越(..)即便有授权也拒绝。
- 选定文件的 read/edit_excel 真能动；写/改前自动备份。
"""
import os

import pytest
from openpyxl import Workbook

from services.agent.context import AgentContext


@pytest.fixture
def library(tmp_path, monkeypatch):
    """把内容库指向临时目录，隔离真实 ~/.billiards-desktop。"""
    lib = tmp_path / "library"
    lib.mkdir()
    monkeypatch.setenv("DESKTOP_LIBRARY_DIR", str(lib))
    return lib


def test_resolve_allows_library_and_denies_outside(library, tmp_path):
    from services.agent import local_tools as lt

    # 库内相对路径 → 放行
    assert lt._resolve("a.txt", AgentContext()) == (library.resolve() / "a.txt")
    # 库外文件、且没选 → 拒绝
    outside = tmp_path / "outside.xlsx"
    with pytest.raises(ValueError):
        lt._resolve(str(outside), AgentContext())


def test_resolve_allows_user_selected_file(library, tmp_path):
    from services.agent import local_tools as lt

    report = tmp_path / "desktop" / "3月报表.xlsx"
    report.parent.mkdir(parents=True)
    report.write_text("x")
    # 用户选了它 → 放行
    ctx = AgentContext(allowed_paths=[str(report)])
    assert lt._resolve(str(report), ctx) == report.resolve()
    # 同目录其它没选的文件 → 仍拒绝（只授权了那个文件，不是整个目录）
    sibling = report.parent / "别的.xlsx"
    with pytest.raises(ValueError):
        lt._resolve(str(sibling), ctx)


def test_resolve_denies_traversal_even_with_grant(library, tmp_path):
    from services.agent import local_tools as lt

    granted_dir = tmp_path / "reports"
    granted_dir.mkdir()
    granted_file = granted_dir / "ok.xlsx"
    granted_file.write_text("o")
    secret = tmp_path / "secret.txt"  # 库外、未授权
    secret.write_text("s")

    # 只授权了 granted_file 这一个文件
    ctx = AgentContext(allowed_paths=[str(granted_file)])
    # 想借 .. 从授权文件所在目录跳到未授权的 secret → 解析后落在 secret，拒绝
    with pytest.raises(ValueError):
        lt._resolve(str(granted_dir / ".." / "secret.txt"), ctx)
    # 直接给未授权的库外文件 → 拒绝
    with pytest.raises(ValueError):
        lt._resolve(str(secret), ctx)


@pytest.mark.asyncio
async def test_edit_excel_on_selected_file_with_backup(library, tmp_path):
    from services.agent import local_tools as lt

    report = tmp_path / "营业额.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["B2"] = 5000
    wb.save(report)

    ctx = AgentContext(allowed_paths=[str(report)])
    # 读得到
    read = await lt.read_file({"path": str(report)}, ctx)
    assert "5000" in read
    # 改得动（选定文件，库外）
    res = await lt.edit_excel({"path": str(report), "changes": [{"cell": "B2", "value": 8600}]}, ctx)
    assert "8600" in res
    # 真落盘了
    from openpyxl import load_workbook
    assert load_workbook(report)["Sheet"]["B2"].value == 8600
    # 备份进了内容库的 .backups（不污染老板原目录）
    backups = list((library / ".backups").glob("*.bak"))
    assert len(backups) == 1


def test_full_disk_access_bypasses_scope(library, tmp_path):
    from services.agent import local_tools as lt

    outside = tmp_path / "anywhere" / "任意.txt"
    outside.parent.mkdir(parents=True)
    # 未开全盘 → 拒
    with pytest.raises(ValueError):
        lt._resolve(str(outside), AgentContext())
    # 开了全盘（高级模式）→ 放行任意路径
    assert lt._resolve(str(outside), AgentContext(full_disk_access=True)) == outside.resolve()


@pytest.mark.asyncio
async def test_edit_excel_denies_unselected_outside(library, tmp_path):
    from services.agent import local_tools as lt

    report = tmp_path / "没选的.xlsx"
    wb = Workbook(); wb.active["A1"] = 1; wb.save(report)
    # 没把它放进 allowed_paths → edit_excel 应因越界抛
    with pytest.raises(ValueError):
        await lt.edit_excel({"path": str(report), "changes": [{"cell": "A1", "value": 2}]}, AgentContext())


# ────────── #M6-2: 不同目录同名文件备份不撞名 + diff 取对原件 ──────────

def test_backup_no_collision_different_dirs(library):
    """两个不同目录下的同名文件，备份应各自独立、互不覆盖。"""
    from services.agent import local_tools as lt

    dir_a = library / "reports" / "a"
    dir_b = library / "reports" / "b"
    dir_a.mkdir(parents=True)
    dir_b.mkdir(parents=True)

    file_a = dir_a / "data.txt"
    file_b = dir_b / "data.txt"
    file_a.write_text("content A")
    file_b.write_text("content B")

    bak_a = lt._backup(file_a)
    bak_b = lt._backup(file_b)

    assert bak_a != bak_b
    from pathlib import Path
    assert Path(bak_a).read_text() == "content A"
    assert Path(bak_b).read_text() == "content B"


def test_backup_diff_correct_for_same_name_files(library):
    """同名不同目录文件各自备份后修改，diff 应取回各自正确的原件。"""
    from services.agent import local_tools as lt

    dir_a = library / "a"
    dir_b = library / "b"
    dir_a.mkdir()
    dir_b.mkdir()

    file_a = dir_a / "report.txt"
    file_b = dir_b / "report.txt"
    file_a.write_text("original A")
    file_b.write_text("original B")

    lt._backup(file_a)
    lt._backup(file_b)
    file_a.write_text("modified A")
    file_b.write_text("modified B")

    diff_a = lt.get_file_backup_diff(str(file_a))
    diff_b = lt.get_file_backup_diff(str(file_b))
    assert diff_a["ok"]
    assert diff_b["ok"]
    assert diff_a["old"] == "original A"
    assert diff_b["old"] == "original B"
    assert diff_a["new"] == "modified A"
    assert diff_b["new"] == "modified B"
