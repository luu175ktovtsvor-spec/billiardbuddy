"""personal / roster 两形态 Excel 渲染单测。"""
import io

from openpyxl import load_workbook

from services.report_excel import render_report

_PERSONAL = {
    "key": "coach_main_daily", "shape": "personal", "title": "{store_name}·主教练日报·{date}",
    "groups": [{"name": "今日动作", "fields": [{"key": "add_wechat", "label": "加微", "unit": "人"}]}],
    "excel": {"layout": "personal"},
}
_ROSTER = {
    "key": "assistant_manager_daily", "shape": "roster", "title": "{store_name}·助教管理日报·{date}",
    "row_label": "助教", "rank_by": "hours",
    "columns": [{"key": "name", "label": "助教"}, {"key": "hours", "label": "陪打时长", "unit": "h"}],
    "excel": {"layout": "roster"},
}


def test_personal_has_today_and_cumulative_columns():
    buf = render_report(
        _PERSONAL,
        {"store_name": "示例", "date": "2026-06-14", "add_wechat": 5, "_cumulative": {"add_wechat": 379}},
        "今日总结：稳。",
    )
    vals = [c.value for row in load_workbook(io.BytesIO(buf)).active.iter_rows() for c in row if c.value is not None]
    assert 5 in vals and 379 in vals          # 今日 + 累计 都在


def test_roster_two_sheets_and_ranked():
    buf = render_report(
        _ROSTER,
        {"store_name": "示例", "date": "2026-06-14", "rows": [{"name": "A", "hours": 3}, {"name": "B", "hours": 5}]},
        "B 领先。",
    )
    wb = load_workbook(io.BytesIO(buf))
    assert "明细" in wb.sheetnames and "排名" in wb.sheetnames
    rank_vals = [c.value for row in wb["排名"].iter_rows() for c in row if c.value is not None]
    assert rank_vals.index("B") < rank_vals.index("A")   # B 排在 A 前
