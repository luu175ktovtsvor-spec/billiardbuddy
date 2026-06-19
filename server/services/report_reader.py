"""经营报表抽取器：把老板从收银系统导出的 Excel 报表读成【决策树用得上的指标】。

输入：一份 .xlsx（典型是「经营数据总表」——含逐日的收入日记账优化表 + 支出表）。
输出：(indicators: dict, summary: str)
  - indicators：算出来的关键指标（团购占比 / 台费占比 / 助教费占比 / 商品费占比 / 客单价 /
    翻台率 / 会员充值vs扣卡 / 新客占比 / 各区开台占比+平均时长 / 房租占比 …），best-effort，
    源列找不到的就不算（记进 missing），绝不抛错把整个诊断带崩。
  - summary：人话可读串，前置拼进诊断的 current_situation，让 diagnostic_logic 决策树据此评估。

设计要点：
- 用 openpyxl(data_only=True) 读公式的计算值。
- 自动挑「含营业额/团购/开台 列」的那张富表（名字含 收入/日记账 优先）。
- 表头在第 2 行（row2，双层表头：row1 大类，row2 细列）；按 fuzzy contains 匹配列名，
  容错不同收银导出的命名差异（如「总营业额」也匹配「营业额合计」）。
- 逐日聚合：营业额/各项费用/单数/团购费 求和；占比/客单价/时长 取均值（按有值的行）。
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


# ── 列名 fuzzy 匹配关键词：每个指标给一组候选片段，表头只要 contains 任一片段即命中 ──
# 顺序敏感：更具体的关键词放前面（避免「营业额」先吃掉「营业额美团」）。
def _norm(s: Any) -> str:
    """表头归一化：转字符串、去空白/换行，便于 contains 比对。"""
    return str(s or "").replace("\n", "").replace(" ", "").replace("\t", "").strip()


def _to_num(v: Any) -> float | None:
    """把单元格值转成数字；带 %、￥、逗号、'h'、'小时' 等也尽量抠出来。None/非数 → None。"""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # 去掉常见单位/符号
    for ch in ("￥", "¥", "%", ",", "，", "元", "次", "人", "单"):
        s = s.replace(ch, "")
    s = s.replace("小时", "").replace("h", "").replace("H", "").strip()
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _has_pct(v: Any) -> bool:
    return isinstance(v, str) and "%" in v


# 指标 → 候选表头片段（fuzzy contains）。注意区分易混项。
_COL_KEYS: dict[str, list[str]] = {
    "revenue":        ["总营业额", "营业额合计", "营业总额"],   # 不含「美团/抖音/各区」
    "orders":         ["总单数", "单数合计", "总订单"],
    "avg_price":      ["平均每单价", "客单价", "单均"],
    "goods_fee":      ["商品费", "商品收入", "商品消费"],
    "table_fee":      ["总开台费", "开台费合计", "台费合计"],
    "assistant_fee":  ["总助教费", "助教费合计", "助教陪打费"],
    "member_recharge": ["会员充值", "充值金额", "充值合计"],
    "member_deduct":  ["会员消费储值扣款", "储值扣款", "扣卡", "扣款"],
    "new_customers":  ["新客人数", "新客数", "新增客户"],
    "turnover_hours": ["平均每桌时长", "每桌平均时长", "翻台率", "平均桌时长"],
    "groupbuy_fee":   ["总团购费", "团购费合计", "团购金额"],
    "groupbuy_ratio": ["团购费占比", "团购占比"],
}

# 各区：名字片段（用于「XX区 开台次数 / 平均时长」的列匹配）
_AREA_NAMES = ["乔氏金腿", "乔氏银腿", "台球包间", "棋牌包间", "独牙自助区", "独牙非自助区"]


def _pick_rich_sheet(wb) -> Any:
    """挑那张『富表』：含 营业额/团购/开台 列的；名字含 收入/日记账 的优先。
    打分：名字命中 + 表头命中的关键列数。"""
    best = None
    best_score = -1
    for ws in wb.worksheets:
        title = _norm(ws.title)
        name_bonus = 0
        if "收入" in title or "日记账" in title:
            name_bonus += 3
        if "优化" in title:
            name_bonus += 1
        # 扫前 3 行表头，数命中多少富表特征列
        hdr_hits = 0
        for r in ws.iter_rows(min_row=1, max_row=3):
            for cell in r:
                h = _norm(cell.value)
                if not h:
                    continue
                if any(k in h for k in ("营业额", "团购", "开台", "助教", "客单价", "充值")):
                    hdr_hits += 1
        score = name_bonus + hdr_hits
        if score > best_score and hdr_hits > 0:
            best, best_score = ws, score
    return best


def _find_header_row(ws) -> int:
    """找表头行：双层表头里『细列』那行（row2 通常是）。取前 4 行里非空单元格最多、且含已知列关键词的那行。"""
    best_row, best_hits = 2, -1
    for r in range(1, min(ws.max_row, 5) + 1):
        hits = 0
        for cell in ws[r]:
            h = _norm(cell.value)
            if not h:
                continue
            for keys in _COL_KEYS.values():
                if any(k in h for k in keys):
                    hits += 1
                    break
        if hits > best_hits:
            best_row, best_hits = r, hits
    return best_row


def _map_columns(ws, header_row: int) -> tuple[dict[str, int], dict[str, dict[str, int]]]:
    """把指标 → 列号(1-based)；返回 (metric_cols, area_cols)。
    area_cols：{区名: {'open': 列, 'dur': 列}}（开台次数 / 平均时长）。"""
    metric_cols: dict[str, int] = {}
    area_cols: dict[str, dict[str, int]] = {}

    cells = list(ws[header_row])
    for cell in cells:
        h = _norm(cell.value)
        if not h:
            continue
        col = cell.column  # 1-based int
        # 普通指标列：fuzzy contains（已占的不覆盖，保前面更具体的）
        for metric, keys in _COL_KEYS.items():
            if metric in metric_cols:
                continue
            # revenue 特判：别被「营业额美团/营业额抖音/各区营收」吃掉
            if metric == "revenue" and ("美团" in h or "抖音" in h or any(a in h for a in _AREA_NAMES)):
                continue
            if any(k in h for k in keys):
                metric_cols[metric] = col
                break
        # 各区开台次数 / 平均时长
        for area in _AREA_NAMES:
            if area in h:
                slot = area_cols.setdefault(area, {})
                if "开台" in h and "次" in h:
                    slot["open"] = col
                elif "时长" in h or "时间" in h:  # 该区平均时长列
                    slot["dur"] = col
    return metric_cols, area_cols


def _data_rows(ws, header_row: int):
    """逐日数据行：表头行之后、首列(或任一关键列)有值的行。"""
    return ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row)


def _col_values(ws, header_row: int, col: int) -> list[float]:
    out: list[float] = []
    for row in _data_rows(ws, header_row):
        if col - 1 < len(row):
            n = _to_num(row[col - 1].value)
            if n is not None:
                out.append(n)
    return out


def _find_rent(wb) -> float | None:
    """从支出表(支出日记账/收支月记账)找『房租』并求和。best-effort。"""
    for ws in wb.worksheets:
        title = _norm(ws.title)
        if not ("支出" in title or "收支" in title or "月记账" in title):
            continue
        total = 0.0
        found = False
        # 房租可能是一列(逐日)，也可能是「项目|金额」两列里一行
        # 先找含「房租」的表头列
        rent_col = None
        for r in range(1, min(ws.max_row, 5) + 1):
            for cell in ws[r]:
                if "房租" in _norm(cell.value):
                    rent_col = cell.column
                    hdr_r = r
                    break
            if rent_col:
                break
        if rent_col:
            for row in ws.iter_rows(min_row=hdr_r + 1, max_row=ws.max_row):
                if rent_col - 1 < len(row):
                    n = _to_num(row[rent_col - 1].value)
                    if n is not None:
                        total += n
                        found = True
            if found:
                return total
        # 兜底：「项目|金额」竖排 —— 找含「房租」的格，取右边一格做金额
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                if "房租" in _norm(cell.value):
                    for j in range(i + 1, len(row)):
                        n = _to_num(row[j].value)
                        if n is not None:
                            return n
        if found:
            return total
    return None


def extract_report_indicators(path: str) -> tuple[dict, str]:
    """读经营报表 → (indicators, summary)。best-effort，绝不抛错（坏文件返回空 indicators + 提示 summary）。"""
    indicators: dict[str, Any] = {}
    missing: list[str] = []

    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
    except Exception as e:  # noqa: BLE001 — 坏文件/非 xlsx，降级，别带崩诊断
        logger.warning("报表读取失败 %s：%s", path, e)
        return {}, f"（报表《{Path(path).name}》没能读出来：{e}。这次按你说的情况诊断。）"

    ws = _pick_rich_sheet(wb)
    if ws is None:
        return {}, f"（报表《{Path(path).name}》里没找到含营业额/团购/开台的经营数据表。这次按你说的情况诊断。）"

    header_row = _find_header_row(ws)
    metric_cols, area_cols = _map_columns(ws, header_row)

    def _sum(metric: str) -> float | None:
        col = metric_cols.get(metric)
        if not col:
            return None
        vals = _col_values(ws, header_row, col)
        return sum(vals) if vals else None

    def _mean(metric: str, pct: bool = False) -> float | None:
        col = metric_cols.get(metric)
        if not col:
            return None
        vals: list[float] = []
        for row in _data_rows(ws, header_row):
            if col - 1 < len(row):
                raw = row[col - 1].value
                n = _to_num(raw)
                if n is None:
                    continue
                # 占比列：Excel 里 0.18 与 "18%" 都可能；统一成「百分数数值」
                if pct and (isinstance(raw, (int, float)) and not _has_pct(raw)) and 0 <= n <= 1:
                    n *= 100
                vals.append(n)
        return sum(vals) / len(vals) if vals else None

    # ── 聚合 ──
    revenue = _sum("revenue")
    orders = _sum("orders")
    goods_fee = _sum("goods_fee")
    table_fee = _sum("table_fee")
    assistant_fee = _sum("assistant_fee")
    groupbuy_fee = _sum("groupbuy_fee")
    member_recharge = _sum("member_recharge")
    member_deduct = _sum("member_deduct")
    new_customers = _sum("new_customers")
    turnover_hours = _mean("turnover_hours")
    groupbuy_ratio_col = _mean("groupbuy_ratio", pct=True)
    avg_price_mean = _mean("avg_price")

    if revenue is not None:
        indicators["total_revenue"] = round(revenue, 2)
    else:
        missing.append("总营业额")

    # 团购占比：优先用「团购费占比」列均值；否则 总团购费/总营业额
    if groupbuy_ratio_col is not None:
        indicators["groupbuy_ratio_pct"] = round(groupbuy_ratio_col, 1)
    elif groupbuy_fee is not None and revenue:
        indicators["groupbuy_ratio_pct"] = round(groupbuy_fee / revenue * 100, 1)
    else:
        missing.append("团购占比")

    # 台费 / 助教费 / 商品费 占比（占总营业额）
    if table_fee is not None and revenue:
        indicators["table_fee_ratio_pct"] = round(table_fee / revenue * 100, 1)
    else:
        missing.append("台费占比")
    if assistant_fee is not None and revenue:
        indicators["assistant_fee_ratio_pct"] = round(assistant_fee / revenue * 100, 1)
    else:
        missing.append("助教费占比")
    if goods_fee is not None and revenue:
        indicators["goods_fee_ratio_pct"] = round(goods_fee / revenue * 100, 1)
    else:
        missing.append("商品费占比")

    # 商品费 vs 助教费（规则4）
    if goods_fee is not None and assistant_fee is not None:
        indicators["goods_fee"] = round(goods_fee, 2)
        indicators["assistant_fee"] = round(assistant_fee, 2)
        indicators["goods_below_assistant"] = goods_fee < assistant_fee

    # 客单价：优先列均值；否则 总营业额/总单数
    if avg_price_mean is not None:
        indicators["avg_price"] = round(avg_price_mean, 1)
    elif revenue is not None and orders:
        indicators["avg_price"] = round(revenue / orders, 1)
    else:
        missing.append("平均客单价")

    # 翻台率（每桌平均时长 h）
    if turnover_hours is not None:
        indicators["turnover_hours"] = round(turnover_hours, 2)
    else:
        missing.append("翻台率/每桌平均时长")

    # 会员充值 vs 储值扣款（规则5 空挂）
    if member_recharge is not None:
        indicators["member_recharge"] = round(member_recharge, 2)
    else:
        missing.append("会员充值")
    if member_deduct is not None:
        indicators["member_deduct"] = round(member_deduct, 2)
    else:
        missing.append("会员储值扣款")
    if member_recharge is not None and member_deduct is not None:
        indicators["recharge_idle"] = member_recharge > 0 and member_deduct < member_recharge * 0.5

    # 新客占比（新客 / 总单数，best-effort）
    if new_customers is not None and orders:
        indicators["new_customer_ratio_pct"] = round(new_customers / orders * 100, 1)
    elif new_customers is not None:
        indicators["new_customers"] = round(new_customers, 0)
    else:
        missing.append("新客占比")

    # 各区开台次数占比 + 平均时长
    area_stats: dict[str, dict[str, float]] = {}
    total_open = 0.0
    for area, slot in area_cols.items():
        oc = slot.get("open")
        opens = sum(_col_values(ws, header_row, oc)) if oc else None
        dc = slot.get("dur")
        durs = _col_values(ws, header_row, dc) if dc else []
        dur_mean = sum(durs) / len(durs) if durs else None
        if opens is not None or dur_mean is not None:
            area_stats[area] = {}
            if opens is not None:
                area_stats[area]["open"] = opens
                total_open += opens
            if dur_mean is not None:
                area_stats[area]["dur"] = round(dur_mean, 2)
    if total_open > 0:
        for area, st in area_stats.items():
            if "open" in st:
                st["open_ratio_pct"] = round(st["open"] / total_open * 100, 1)
    if area_stats:
        indicators["areas"] = area_stats
    else:
        missing.append("各区开台与平均时长")

    # 房租占比（从支出表）
    rent = _find_rent(wb)
    if rent is not None and revenue:
        indicators["rent_ratio_pct"] = round(rent / revenue * 100, 1)
        indicators["rent"] = round(rent, 2)
    else:
        missing.append("房租占比")

    if missing:
        indicators["missing"] = missing

    summary = _build_summary(indicators, Path(path).name, missing)
    return indicators, summary


def _build_summary(ind: dict, filename: str, missing: list[str]) -> str:
    """把指标拼成给决策树读的人话串（带健康/预警/紧急的标尺）。"""
    L: list[str] = [f"【报表关键指标（系统从你的报表《{filename}》自动算出）】"]

    if "total_revenue" in ind:
        L.append(f"- 总营业额：{ind['total_revenue']:g}")
    if "groupbuy_ratio_pct" in ind:
        L.append(f"- 团购占比：{ind['groupbuy_ratio_pct']:g}%（健康<20 / 预警>40 / 紧急>70）")
    # 台费/助教费/商品费占比
    fee_bits = []
    if "table_fee_ratio_pct" in ind:
        fee_bits.append(f"台费 {ind['table_fee_ratio_pct']:g}%（正常60-70）")
    if "assistant_fee_ratio_pct" in ind:
        fee_bits.append(f"助教费 {ind['assistant_fee_ratio_pct']:g}%")
    if "goods_fee_ratio_pct" in ind:
        fee_bits.append(f"商品费 {ind['goods_fee_ratio_pct']:g}%")
    if fee_bits:
        L.append("- 收入结构占比：" + "、".join(fee_bits))
    if ind.get("goods_below_assistant") is True:
        L.append("  · 商品费 < 助教费（命中『推品能力不足』规则4）")
    if "avg_price" in ind:
        L.append(f"- 平均客单价：{ind['avg_price']:g}")
    if "turnover_hours" in ind:
        L.append(f"- 翻台率（每桌平均时长）：{ind['turnover_hours']:g} 小时（4h 及格 / 6h 优秀）")
    if "member_recharge" in ind or "member_deduct" in ind:
        r = ind.get("member_recharge", "?")
        d = ind.get("member_deduct", "?")
        tail = "（充值远大于扣卡=会员卡在『空挂』，命中规则5）" if ind.get("recharge_idle") else ""
        L.append(f"- 会员充值 {r if r == '?' else f'{r:g}'} / 储值扣卡 {d if d == '?' else f'{d:g}'}{tail}")
    if "new_customer_ratio_pct" in ind:
        L.append(f"- 新客占比：{ind['new_customer_ratio_pct']:g}%")
    elif "new_customers" in ind:
        L.append(f"- 新客人数：{ind['new_customers']:g}")
    if ind.get("areas"):
        parts = []
        for area, st in ind["areas"].items():
            seg = area
            if "open_ratio_pct" in st:
                seg += f" 开台占比{st['open_ratio_pct']:g}%"
            elif "open" in st:
                seg += f" 开台{st['open']:g}次"
            if "dur" in st:
                seg += f" 均时长{st['dur']:g}h"
            parts.append(seg)
        L.append("- 各区开台与平均时长：" + "；".join(parts) + "（开台率<30紧急、平均时长<1.5h预警）")
    if "rent_ratio_pct" in ind:
        L.append(f"- 房租占比：{ind['rent_ratio_pct']:g}%（建议不超过营业额30%）")

    if missing:
        L.append(f"- （报表里没找到的指标，按你说的情况补：{('、'.join(missing))}）")

    return "\n".join(L)
