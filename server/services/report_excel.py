"""报表 Excel 渲染（纯函数）。render_report(schema, data, narrative) -> bytes。

按 schema["shape"] 分派 flat / personal / roster。设计为纯函数：入参是 dict，
出参是 bytes，不碰 DB，便于单测。
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_TITLE_FONT = Font(bold=True, size=14)
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="007AFF")  # iOS 系统蓝，对齐设计系统
_WRAP = Alignment(wrap_text=True, vertical="top")


def _title(schema: dict, data: dict) -> str:
    t = schema.get("title", schema["key"])
    for k, v in data.items():
        t = t.replace("{" + k + "}", str(v))
    return t


def render_report(schema: dict, data: dict, narrative: str) -> bytes:
    shape = schema["shape"]
    if shape == "flat":
        wb = _render_flat(schema, data, narrative)
    elif shape == "personal":
        wb = _render_personal(schema, data, narrative)
    elif shape == "roster":
        wb = _render_roster(schema, data, narrative)
    else:
        raise ValueError(f"未知 shape: {shape}")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_flat(schema: dict, data: dict, narrative: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "日报"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    r = 1
    ws.cell(r, 1, _title(schema, data)).font = _TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 2
    for group in schema["groups"]:
        c = ws.cell(r, 1, group["name"])
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        ws.cell(r, 2, "").fill = _HEAD_FILL
        r += 1
        for f in group["fields"]:
            ws.cell(r, 1, f["label"])
            val = data.get(f["key"])
            unit = f.get("unit", "")
            ws.cell(r, 2, val if unit in ("", None) or val is None else f"{val}{unit}")
            r += 1
        r += 1
    c = ws.cell(r, 1, "AI 叙事")
    c.font = _HEAD_FONT
    c.fill = _HEAD_FILL
    ws.cell(r, 2, "").fill = _HEAD_FILL
    r += 1
    cell = ws.cell(r, 1, narrative or "")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell.alignment = _WRAP
    ws.row_dimensions[r].height = 120
    return wb


def _render_personal(schema: dict, data: dict, narrative: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "日报"
    ws.column_dimensions["A"].width = 18
    for col in ("B", "C"):
        ws.column_dimensions[col].width = 14
    cum = data.get("_cumulative", {})
    r = 1
    ws.cell(r, 1, _title(schema, data)).font = _TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 2
    for col, txt in ((1, "项目"), (2, "今日"), (3, "本月累计")):
        c = ws.cell(r, col, txt)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
    r += 1
    for group in schema["groups"]:
        for f in group["fields"]:
            ws.cell(r, 1, f["label"])
            ws.cell(r, 2, data.get(f["key"]))
            ws.cell(r, 3, cum.get(f["key"], ""))
            r += 1
    r += 1
    c = ws.cell(r, 1, "AI 叙事")
    c.font = _HEAD_FONT
    c.fill = _HEAD_FILL
    r += 1
    cell = ws.cell(r, 1, narrative or "")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    cell.alignment = _WRAP
    ws.row_dimensions[r].height = 100
    return wb


def _render_roster(schema: dict, data: dict, narrative: str) -> Workbook:
    from services.report_service import rank_roster

    cols = schema["columns"]
    rows = data.get("rows", [])
    wb = Workbook()
    detail = wb.active
    detail.title = "明细"
    detail.cell(1, 1, _title(schema, data)).font = _TITLE_FONT
    detail.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    for j, col in enumerate(cols, 1):
        c = detail.cell(3, j, col["label"])
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
    for i, row in enumerate(rows, 4):
        for j, col in enumerate(cols, 1):
            detail.cell(i, j, row.get(col["key"]))

    id_key = cols[0]["key"]
    rank_by = schema.get("rank_by") or (cols[1]["key"] if len(cols) > 1 else id_key)
    rank_label = next((c["label"] for c in cols if c["key"] == rank_by), rank_by)
    rank_sheet = wb.create_sheet("排名")
    ranked = rank_roster([dict(x) for x in rows], rank_by)
    for j, h in enumerate(("排名", schema.get("row_label", "成员"), rank_label), 1):
        c = rank_sheet.cell(1, j, h)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
    for i, row in enumerate(ranked, 2):
        rank_sheet.cell(i, 1, row["rank"])
        rank_sheet.cell(i, 2, row.get(id_key))
        rank_sheet.cell(i, 3, row.get(rank_by))

    note = wb.create_sheet("播报")
    note.cell(1, 1, narrative or "").alignment = _WRAP
    note.column_dimensions["A"].width = 60
    return wb
